from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import Optional
from datetime import datetime, timezone
import uuid
from io import BytesIO

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from sqlalchemy import select, update, delete
from sqlalchemy.ext.asyncio import AsyncSession

from database import Client, serialize_doc, get_db
from routes.auth import get_current_user

router = APIRouter(prefix="/clients", tags=["Clients"])


class ClientCreate(BaseModel):
    name: str
    type: str
    gstin: Optional[str] = None
    email: EmailStr
    phone: str
    address: str
    city: str
    state: str
    pincode: str


@router.get("")
async def get_clients(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]).limit(5000))
    return [serialize_doc(c) for c in result.scalars().all()]


@router.post("")
async def create_client(client_data: ClientCreate, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    client_id = f"client_{uuid.uuid4().hex[:12]}"
    new_client = Client(
        client_id=client_id,
        tenant_id=user["tenant_id"],
        outstanding_balance=0.0,
        created_at=datetime.now(timezone.utc),
        **client_data.model_dump()
    )
    db.add(new_client)
    await db.commit()
    await db.refresh(new_client)
    return serialize_doc(new_client)


@router.post("/upload-excel")
async def upload_clients_excel(file: UploadFile = File(...), user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    try:
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        expected_fields = {
            'name': ['name', 'client name', 'company name'],
            'type': ['type', 'client type'],
            'gstin': ['gstin', 'gst number', 'gst no', 'gst'],
            'email': ['email', 'email address', 'e-mail'],
            'phone': ['phone', 'mobile', 'contact', 'phone number'],
            'address': ['address', 'street address'],
            'city': ['city'], 'state': ['state'],
            'pincode': ['pincode', 'pin code', 'zip', 'zip code', 'postal code']
        }
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                for field, aliases in expected_fields.items():
                    if header_lower in aliases:
                        header_map[field] = idx
                        break

        required = ['name', 'email', 'phone', 'address', 'city', 'state', 'pincode']
        missing = [f for f in required if f not in header_map]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

        created_count = 0; skipped_count = 0; errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = str(row[header_map['name']]).strip() if row[header_map['name']] else None
                if not name:
                    continue
                email = str(row[header_map['email']]).strip() if row[header_map['email']] else None
                if not email or '@' not in email:
                    errors.append(f"Row {row_idx}: Invalid email"); skipped_count += 1; continue

                existing = await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"], Client.email == email))
                if existing.scalar_one_or_none():
                    skipped_count += 1; continue

                db.add(Client(
                    client_id=f"client_{uuid.uuid4().hex[:12]}",
                    tenant_id=user["tenant_id"],
                    name=name,
                    type=str(row[header_map.get('type', 0)] or 'B2B').strip() if header_map.get('type') is not None else 'B2B',
                    gstin=str(row[header_map['gstin']]).strip() if header_map.get('gstin') is not None and row[header_map['gstin']] else None,
                    email=email,
                    phone=str(row[header_map['phone']]).strip() if row[header_map['phone']] else '',
                    address=str(row[header_map['address']]).strip() if row[header_map['address']] else '',
                    city=str(row[header_map['city']]).strip() if row[header_map['city']] else '',
                    state=str(row[header_map['state']]).strip() if row[header_map['state']] else '',
                    pincode=str(row[header_map['pincode']]).strip() if row[header_map['pincode']] else '',
                    outstanding_balance=0.0,
                    created_at=datetime.now(timezone.utc)
                ))
                created_count += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}"); skipped_count += 1

        await db.commit()
        return {"message": f"Upload complete: {created_count} clients created, {skipped_count} skipped",
                "created": created_count, "skipped": skipped_count, "errors": errors[:10]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


@router.put("/{client_id}")
async def update_client(client_id: str, client_data: ClientCreate, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Client).where(Client.client_id == client_id, Client.tenant_id == user["tenant_id"]))
    client = result.scalar_one_or_none()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    for k, v in client_data.model_dump().items():
        setattr(client, k, v)
    await db.commit()
    await db.refresh(client)
    return serialize_doc(client)


@router.delete("/{client_id}")
async def delete_client(client_id: str, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(delete(Client).where(Client.client_id == client_id, Client.tenant_id == user["tenant_id"]))
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    await db.commit()
    return {"message": "Client deleted successfully"}


@router.get("/template/download")
async def download_client_template(user: dict = Depends(get_current_user)):
    wb = Workbook(); ws = wb.active; ws.title = "Clients Template"
    headers = ["Name", "Type", "GSTIN", "Email", "Phone", "Address", "City", "State", "Pincode"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    sample = ["ABC Company Pvt Ltd", "B2B", "29ABCDE1234F1Z5", "contact@abccompany.com",
              "9876543210", "123, MG Road", "Bangalore", "Karnataka", "560001"]
    for col_num, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_num, value=val).border = border
    for i, w in enumerate([25, 10, 20, 30, 15, 30, 15, 15, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=client_upload_template.xlsx"})
