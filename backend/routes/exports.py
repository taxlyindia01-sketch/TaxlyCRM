from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database import Invoice, Payment, Client, Estimate, CreditNote, Business, serialize_doc, get_db
from routes.auth import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])


def style_worksheet(ws, header_row=1):
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[header_row]:
        if not isinstance(cell, MergedCell):
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    for row in ws.iter_rows(min_row=header_row+1):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
    for col_idx in range(1, ws.max_column+1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row+1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
        ws.column_dimensions[col_letter].width = min(max(max_length+2, 10), 50)


@router.get("/invoices")
async def export_invoices(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None),
                           user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    query = select(Invoice).where(Invoice.tenant_id == user["tenant_id"])
    if start_date and end_date:
        query = query.where(Invoice.invoice_date >= start_date, Invoice.invoice_date <= end_date)
    invoices = (await db.execute(query)).scalars().all()
    clients_raw = (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()
    clients = {c.client_id: c for c in clients_raw}

    wb = Workbook()
    ws_sum = wb.active; ws_sum.title = "Invoices Summary"
    sum_headers = ["Invoice #", "Date", "Client", "GSTIN", "State", "Diary No.", "Ref No.", "Subtotal", "CGST", "SGST", "IGST", "Total", "Currency", "Status", "Paid", "Outstanding"]
    for col, h in enumerate(sum_headers, 1): ws_sum.cell(row=1, column=col, value=h)
    ws_items = wb.create_sheet("Invoices Items")
    item_headers = ["Invoice #", "Date", "Client", "GSTIN", "State", "Description", "HSN/SAC", "Qty", "Rate", "GST%", "Amount", "Currency"]
    for col, h in enumerate(item_headers, 1): ws_items.cell(row=1, column=col, value=h)

    row_s = 2; row_i = 2
    for inv in invoices:
        c = clients.get(inv.client_id)
        cname = c.name if c else inv.client_name or 'Unknown'
        cgstin = c.gstin if c else inv.client_gstin or ''
        ws_sum.cell(row=row_s, column=1, value=inv.invoice_number); ws_sum.cell(row=row_s, column=2, value=inv.invoice_date)
        ws_sum.cell(row=row_s, column=3, value=cname); ws_sum.cell(row=row_s, column=4, value=cgstin)
        ws_sum.cell(row=row_s, column=5, value=inv.client_state or ''); ws_sum.cell(row=row_s, column=6, value=inv.diary_no or '')
        ws_sum.cell(row=row_s, column=7, value=inv.ref_no or ''); ws_sum.cell(row=row_s, column=8, value=inv.subtotal or 0)
        ws_sum.cell(row=row_s, column=9, value=inv.cgst or 0); ws_sum.cell(row=row_s, column=10, value=inv.sgst or 0)
        ws_sum.cell(row=row_s, column=11, value=inv.igst or 0); ws_sum.cell(row=row_s, column=12, value=inv.total or 0)
        ws_sum.cell(row=row_s, column=13, value=inv.currency); ws_sum.cell(row=row_s, column=14, value=inv.status)
        ws_sum.cell(row=row_s, column=15, value=inv.advance_paid or 0); ws_sum.cell(row=row_s, column=16, value=inv.outstanding or 0)
        row_s += 1
        for item in (inv.items or []):
            ws_items.cell(row=row_i, column=1, value=inv.invoice_number); ws_items.cell(row=row_i, column=2, value=inv.invoice_date)
            ws_items.cell(row=row_i, column=3, value=cname); ws_items.cell(row=row_i, column=4, value=cgstin)
            ws_items.cell(row=row_i, column=5, value=inv.client_state or '')
            ws_items.cell(row=row_i, column=6, value=item.get('description',''))
            ws_items.cell(row=row_i, column=7, value=item.get('hsn_sac_code',''))
            ws_items.cell(row=row_i, column=8, value=item.get('quantity',0))
            ws_items.cell(row=row_i, column=9, value=item.get('rate',0))
            ws_items.cell(row=row_i, column=10, value=item.get('gst_rate',0))
            ws_items.cell(row=row_i, column=11, value=item.get('amount',0))
            ws_items.cell(row=row_i, column=12, value=inv.currency)
            row_i += 1

    style_worksheet(ws_sum); style_worksheet(ws_items)
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=invoices.xlsx"})


@router.get("/payments")
async def export_payments(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None),
                           user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    query = select(Payment).where(Payment.tenant_id == user["tenant_id"])
    if start_date and end_date:
        query = query.where(Payment.payment_date >= start_date, Payment.payment_date <= end_date)
    payments = (await db.execute(query)).scalars().all()
    invoices = {i.invoice_id: i for i in (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"]))).scalars().all()}
    clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}

    wb = Workbook(); ws = wb.active; ws.title = "Payments"
    headers = ["Payment Date", "Invoice #", "Client", "Amount", "Payment Mode", "Reference #", "Notes"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    for row_num, p in enumerate(payments, 2):
        inv = invoices.get(p.invoice_id); cli = clients.get(p.client_id)
        ws.cell(row=row_num, column=1, value=p.payment_date)
        ws.cell(row=row_num, column=2, value=inv.invoice_number if inv else "N/A")
        ws.cell(row=row_num, column=3, value=cli.name if cli else "Unknown")
        ws.cell(row=row_num, column=4, value=p.amount)
        ws.cell(row=row_num, column=5, value=p.payment_mode)
        ws.cell(row=row_num, column=6, value=p.reference_number or '')
        ws.cell(row=row_num, column=7, value=p.notes or '')
    style_worksheet(ws)
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=payments.xlsx"})


@router.get("/clients")
async def export_clients(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    clients = (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()
    wb = Workbook(); ws = wb.active; ws.title = "Clients"
    headers = ["Client Name", "Type", "GSTIN", "Email", "Phone", "City", "State", "Outstanding Balance"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    for row_num, c in enumerate(clients, 2):
        ws.cell(row=row_num, column=1, value=c.name); ws.cell(row=row_num, column=2, value=c.type)
        ws.cell(row=row_num, column=3, value=c.gstin or ''); ws.cell(row=row_num, column=4, value=c.email)
        ws.cell(row=row_num, column=5, value=c.phone); ws.cell(row=row_num, column=6, value=c.city)
        ws.cell(row=row_num, column=7, value=c.state); ws.cell(row=row_num, column=8, value=c.outstanding_balance or 0)
    style_worksheet(ws)
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=clients.xlsx"})


@router.get("/outstanding")
async def export_outstanding(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"], Invoice.outstanding > 0))).scalars().all()
    clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}
    wb = Workbook(); ws = wb.active; ws.title = "Outstanding"
    headers = ["Invoice #", "Date", "Client", "Total", "Paid", "Outstanding", "Status"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    for row_num, inv in enumerate(invoices, 2):
        c = clients.get(inv.client_id)
        ws.cell(row=row_num, column=1, value=inv.invoice_number); ws.cell(row=row_num, column=2, value=inv.invoice_date)
        ws.cell(row=row_num, column=3, value=c.name if c else "Unknown")
        ws.cell(row=row_num, column=4, value=inv.total or 0); ws.cell(row=row_num, column=5, value=inv.advance_paid or 0)
        ws.cell(row=row_num, column=6, value=inv.outstanding or 0); ws.cell(row=row_num, column=7, value=inv.status)
    style_worksheet(ws)
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=outstanding.xlsx"})
