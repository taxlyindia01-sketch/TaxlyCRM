from fastapi import APIRouter, HTTPException, Response, Request, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone, timedelta
import uuid
import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from sqlalchemy import select, update, delete
from sqlalchemy.ext.asyncio import AsyncSession

from database import (
    User, AdminSession, Business, Client, Invoice, Estimate,
    CreditNote, Payment, Advance, MasterData,
    serialize_doc, get_db
)

router = APIRouter(prefix="/admin", tags=["Admin"])

ADMIN_USERNAME = (os.environ.get("ADMIN_USERNAME") or
                  os.environ.get("TAXLY_ADMIN_USERNAME") or "Taxlyindia")
_raw_password = os.environ.get("ADMIN_PASSWORD", "@Gtm025@")


class AdminLogin(BaseModel):
    username: str
    password: str


class TenantUpdate(BaseModel):
    is_active: bool


class TenantApproval(BaseModel):
    is_approved: bool
    rejection_reason: Optional[str] = None


async def check_admin(request: Request, db: AsyncSession = Depends(get_db)):
    session_token = request.cookies.get("admin_session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    if not session_token:
        raise HTTPException(status_code=401, detail="Admin not authenticated")

    result = await db.execute(select(AdminSession).where(AdminSession.session_token == session_token))
    session_doc = result.scalar_one_or_none()
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid admin session")
    return True


@router.post("/login")
async def admin_login(login_data: AdminLogin, response: Response, db: AsyncSession = Depends(get_db)):
    import bcrypt as _bcrypt
    _hash = os.environ.get("ADMIN_PASSWORD_HASH", "")
    if login_data.username != ADMIN_USERNAME:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if _hash:
        try: _ok = _bcrypt.checkpw(login_data.password.encode(), _hash.encode())
        except: _ok = False
        if not _ok:
            raise HTTPException(status_code=401, detail="Invalid credentials")
    elif login_data.password != _raw_password:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    session_token = f"admin_session_{uuid.uuid4().hex}"
    db.add(AdminSession(
        session_token=session_token,
        username=ADMIN_USERNAME,
        expires_at=datetime.now(timezone.utc) + timedelta(days=1),
        created_at=datetime.now(timezone.utc)
    ))
    await db.commit()

    _is_https = str(response.headers.get("x-forwarded-proto","http")).startswith("https")
    response.set_cookie(key="admin_session_token", value=session_token, httponly=True,
                        secure=_is_https, samesite="none" if _is_https else "lax",
                        max_age=86400, path="/")
    return {"message": "Admin logged in successfully", "session_token": session_token}


@router.get("/tenants")
async def get_all_tenants(admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User))
    return [serialize_doc(u) for u in result.scalars().all()]


@router.get("/tenants/pending")
async def get_pending_approvals(admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    now = datetime.now(timezone.utc)
    result = await db.execute(select(User).where(User.is_approved == False))
    users = result.scalars().all()

    pending = []
    for user in users:
        user_data = serialize_doc(user)
        demo_expires_at = user.demo_expires_at
        if demo_expires_at:
            if demo_expires_at.tzinfo is None:
                demo_expires_at = demo_expires_at.replace(tzinfo=timezone.utc)
            days_remaining = (demo_expires_at - now).days
            user_data['demo_days_remaining'] = max(0, days_remaining)
            user_data['demo_expired'] = days_remaining < 0
        else:
            user_data['demo_days_remaining'] = 0
            user_data['demo_expired'] = True
        pending.append(user_data)

    pending.sort(key=lambda x: (not x.get('demo_expired', True), x.get('demo_days_remaining', 0)))
    return pending


@router.put("/tenants/{user_id}")
async def update_tenant_status(user_id: str, tenant_update: TenantUpdate,
                                admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.user_id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Tenant not found")
    user.is_active = tenant_update.is_active
    await db.commit()
    return {"message": "Tenant status updated successfully"}


@router.put("/tenants/{user_id}/approve")
async def approve_tenant(user_id: str, approval: TenantApproval,
                          admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.user_id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Tenant not found")

    user.is_approved = approval.is_approved
    user.approval_status = "approved" if approval.is_approved else "rejected"
    user.approved_by = ADMIN_USERNAME
    if approval.is_approved:
        user.approved_at = datetime.now(timezone.utc)
    if not approval.is_approved and approval.rejection_reason:
        user.rejection_reason = approval.rejection_reason
    await db.commit()

    action = "approved" if approval.is_approved else "rejected"
    return {"message": f"Tenant {action} successfully"}


@router.delete("/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.tenant_id == tenant_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Tenant not found")

    for model in [User, Business, Client, Invoice, Estimate, CreditNote, Payment, Advance, MasterData]:
        await db.execute(delete(model).where(model.tenant_id == tenant_id))
    await db.commit()
    return {"message": f"Tenant {tenant_id} and all associated data deleted successfully"}


@router.get("/tenants/{tenant_id}/export")
async def export_tenant_database(tenant_id: str, admin: bool = Depends(check_admin), db: AsyncSession = Depends(get_db)):
    user_res = await db.execute(select(User).where(User.tenant_id == tenant_id))
    user = user_res.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Tenant not found")

    business = (await db.execute(select(Business).where(Business.tenant_id == tenant_id))).scalar_one_or_none()
    clients   = (await db.execute(select(Client).where(Client.tenant_id == tenant_id))).scalars().all()
    invoices  = (await db.execute(select(Invoice).where(Invoice.tenant_id == tenant_id))).scalars().all()
    estimates = (await db.execute(select(Estimate).where(Estimate.tenant_id == tenant_id))).scalars().all()
    credit_notes = (await db.execute(select(CreditNote).where(CreditNote.tenant_id == tenant_id))).scalars().all()
    payments  = (await db.execute(select(Payment).where(Payment.tenant_id == tenant_id))).scalars().all()
    advances  = (await db.execute(select(Advance).where(Advance.tenant_id == tenant_id))).scalars().all()
    master    = (await db.execute(select(MasterData).where(MasterData.tenant_id == tenant_id))).scalars().all()

    wb = Workbook()
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    def add_sheet(wb, title, headers, rows):
        ws = wb.create_sheet(title=title)
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row_num, row_data in enumerate(rows, 2):
            for col_num, val in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=str(val) if val else "").border = border
        for column in ws.columns:
            max_len = max((len(str(c.value)) for c in column if c.value), default=0)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)

    ws = wb.active; ws.title = "Tenant Info"
    rows = [["Field", "Value"], ["Tenant ID", tenant_id], ["User Name", user.name],
            ["Email", user.email], ["Status", "Active" if user.is_active else "Disabled"],
            ["Created", str(user.created_at)]]
    if business:
        rows += [["Business Name", business.business_name], ["GSTIN", business.gstin],
                 ["Business Email", business.email], ["Phone", business.phone],
                 ["Address", business.address], ["City", business.city], ["State", business.state]]
    for row_num, row_data in enumerate(rows, 1):
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            if row_num == 1:
                cell.fill = header_fill; cell.font = header_font
            cell.border = border

    add_sheet(wb, "Clients",
        ["Client ID", "Name", "Type", "GSTIN", "Email", "Phone", "City", "State", "Outstanding"],
        [[c.client_id, c.name, c.type, c.gstin, c.email, c.phone, c.city, c.state, c.outstanding_balance] for c in clients])
    add_sheet(wb, "Invoices",
        ["Invoice #", "Date", "Due Date", "Client ID", "Subtotal", "CGST", "SGST", "IGST", "Total", "Currency", "Status", "Paid", "Outstanding"],
        [[i.invoice_number, i.invoice_date, i.due_date, i.client_id, i.subtotal, i.cgst, i.sgst, i.igst, i.total, i.currency, i.status, i.advance_paid, i.outstanding] for i in invoices])
    add_sheet(wb, "Estimates",
        ["Estimate #", "Date", "Valid Until", "Client ID", "Total", "Currency", "Status"],
        [[e.estimate_number, e.estimate_date, e.valid_until, e.client_id, e.total, e.currency, e.status] for e in estimates])
    add_sheet(wb, "Credit Notes",
        ["Credit Note #", "Date", "Invoice ID", "Client ID", "Total", "Currency", "Reason"],
        [[cn.credit_note_number, cn.credit_date, cn.invoice_id, cn.client_id, cn.total, cn.currency, cn.reason] for cn in credit_notes])
    add_sheet(wb, "Payments",
        ["Payment ID", "Date", "Invoice ID", "Client ID", "Amount", "Mode", "Reference"],
        [[p.payment_id, p.payment_date, p.invoice_id, p.client_id, p.amount, p.payment_mode, p.reference_number] for p in payments])
    add_sheet(wb, "Advances",
        ["Advance ID", "Date", "Client ID", "Amount", "Mode", "Status", "Reference"],
        [[a.advance_id, a.payment_date, a.client_id, a.amount, a.payment_mode, a.status, a.reference_number] for a in advances])
    add_sheet(wb, "Master Data",
        ["Master ID", "Type", "Value", "GST Rate"],
        [[m.master_id, m.type, m.value, m.gst_rate] for m in master])

    output = BytesIO(); wb.save(output); output.seek(0)
    safe_name = user.name.replace(' ', '_')[:20]
    filename = f"tenant_data_{safe_name}_{tenant_id[:8]}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/logout")
async def admin_logout(request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    session_token = request.cookies.get("admin_session_token")
    if session_token:
        await db.execute(delete(AdminSession).where(AdminSession.session_token == session_token))
        await db.commit()
    response.delete_cookie("admin_session_token", path="/")
    return {"message": "Admin logged out successfully"}
