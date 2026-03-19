from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from io import BytesIO
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database import Invoice, Payment, Client, Estimate, CreditNote, Advance, Business, serialize_doc, get_db
from routes.auth import get_current_user

def add_totals_row(ws, numeric_cols, start_row=2, label_col=1, label="TOTAL"):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    if ws.max_row < start_row: return
    total_row = ws.max_row + 1
    total_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    total_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='medium'), bottom=Side(style='medium'))
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        if col_idx == label_col:
            cell.value = label
        elif col_idx in numeric_cols:
            col_vals = [ws.cell(row=r, column=col_idx).value or 0
                        for r in range(start_row, total_row)
                        if isinstance(ws.cell(row=r, column=col_idx).value, (int, float))]
            cell.value = round(sum(col_vals), 2)
        cell.fill = total_fill; cell.font = total_font
        cell.alignment = Alignment(horizontal='right' if col_idx in numeric_cols else 'center')
        cell.border = border

router = APIRouter(prefix="/reports", tags=["Reports"])

def _inr(amount, currency, exchange_rate):
    """Convert amount to INR. If currency is INR or no rate, return as-is."""
    if not currency or currency == 'INR' or not exchange_rate:
        return round(amount or 0, 2)
    return round((amount or 0) * exchange_rate, 2)



def style_header(ws, row=1):
    fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws[row]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def auto_cols(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)


@router.get("/gstr1")
async def get_gstr1_report(start_date: str = Query(...), end_date: str = Query(...),
                             user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    try:
        invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"],
                                                             Invoice.invoice_date >= start_date, Invoice.invoice_date <= end_date))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}

        b2b, b2c = [], []
        for inv in invoices:
            c = clients.get(inv.client_id, None)
            gstin = c.gstin if c else inv.client_gstin
            cur = inv.currency or 'INR'
            exr = inv.exchange_rate or 1
            row = {"invoice_number": inv.invoice_number, "invoice_date": inv.invoice_date,
                   "client_name": c.name if c else inv.client_name or "Unknown",
                   "client_gstin": gstin or "", "place_of_supply": inv.client_state or (c.state if c else ""),
                   "currency": cur, "currency_amount": inv.subtotal or 0, "exchange_rate": exr,
                   "taxable_value": _inr(inv.subtotal or 0, cur, exr),
                   "cgst": _inr(inv.cgst or 0, cur, exr), "sgst": _inr(inv.sgst or 0, cur, exr),
                   "igst": _inr(inv.igst or 0, cur, exr),
                   "total_tax": _inr((inv.cgst or 0)+(inv.sgst or 0)+(inv.igst or 0), cur, exr),
                   "invoice_value": _inr(inv.total or 0, cur, exr)}
            (b2b if gstin else b2c).append(row)

        wb = Workbook()
        ws_b2b = wb.active; ws_b2b.title = "B2B Invoices"
        b2b_headers = ["Invoice No", "Invoice Date", "Client Name", "GSTIN", "Place of Supply",
                       "Currency", "Orig. Amount", "Exch. Rate", "Taxable (INR)", "CGST (INR)", "SGST (INR)", "IGST (INR)", "Total Tax (INR)", "Invoice Value (INR)"]
        for col, h in enumerate(b2b_headers, 1): ws_b2b.cell(row=1, column=col, value=h)
        for r, d in enumerate(b2b, 2):
            for col, k in enumerate(["invoice_number","invoice_date","client_name","client_gstin","place_of_supply","currency","currency_amount","exchange_rate","taxable_value","cgst","sgst","igst","total_tax","invoice_value"], 1):
                ws_b2b.cell(row=r, column=col, value=d[k])
        add_totals_row(ws_b2b, {7,9,10,11,12,13,14})
        style_header(ws_b2b); auto_cols(ws_b2b)

        ws_b2c = wb.create_sheet("B2C Invoices")
        b2c_headers = ["Invoice No", "Invoice Date", "Client Name", "Place of Supply", "Currency", "Orig. Amount", "Exch. Rate", "Taxable (INR)", "CGST (INR)", "SGST (INR)", "IGST (INR)", "Total Tax (INR)", "Invoice Value (INR)"]
        for col, h in enumerate(b2c_headers, 1): ws_b2c.cell(row=1, column=col, value=h)
        for r, d in enumerate(b2c, 2):
            for col, k in enumerate(["invoice_number","invoice_date","client_name","place_of_supply","currency","currency_amount","exchange_rate","taxable_value","cgst","sgst","igst","total_tax","invoice_value"], 1):
                ws_b2c.cell(row=r, column=col, value=d[k])
        add_totals_row(ws_b2c, {6,8,9,10,11,12,13})
        style_header(ws_b2c); auto_cols(ws_b2c)

        # Credit Notes sheet (CDNR - Credit/Debit Note Register)
        credit_notes_list = (await db.execute(select(CreditNote).where(
            CreditNote.tenant_id == user["tenant_id"],
            CreditNote.credit_date >= start_date, CreditNote.credit_date <= end_date
        ))).scalars().all()

        ws_cdnr = wb.create_sheet("CDNR - Credit Notes")
        cdnr_headers = ["Credit Note No", "Credit Note Date", "Client Name", "Client GSTIN",
                        "Place of Supply", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax", "Credit Note Value"]
        for col, h in enumerate(cdnr_headers, 1): ws_cdnr.cell(row=1, column=col, value=h)
        for r, cn in enumerate(credit_notes_list, 2):
            cli = clients.get(cn.client_id)
            ws_cdnr.cell(row=r, column=1, value=cn.credit_note_number)
            ws_cdnr.cell(row=r, column=2, value=cn.credit_date)
            ws_cdnr.cell(row=r, column=3, value=cli.name if cli else "Unknown")
            ws_cdnr.cell(row=r, column=4, value=cli.gstin if cli else "")
            ws_cdnr.cell(row=r, column=5, value=cn.client_state or (cli.state if cli else ""))
            ws_cdnr.cell(row=r, column=6, value=cn.subtotal or 0)
            ws_cdnr.cell(row=r, column=7, value=cn.cgst or 0)
            ws_cdnr.cell(row=r, column=8, value=cn.sgst or 0)
            ws_cdnr.cell(row=r, column=9, value=cn.igst or 0)
            total_tax = (cn.cgst or 0) + (cn.sgst or 0) + (cn.igst or 0)
            ws_cdnr.cell(row=r, column=10, value=total_tax)
            ws_cdnr.cell(row=r, column=11, value=cn.total or 0)
        add_totals_row(ws_cdnr, {6,7,8,9,10,11})
        style_header(ws_cdnr); auto_cols(ws_cdnr)

        output = BytesIO(); wb.save(output); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition": f"attachment; filename=GSTR1_{start_date}_to_{end_date}.xlsx"})
    except Exception as _exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report error: {_exc}")

@router.get("/payment-register")
async def get_payment_register(start_date: str = Query(...), end_date: str = Query(...),
                                 user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    try:
        payments = (await db.execute(select(Payment).where(Payment.tenant_id == user["tenant_id"],
                                                             Payment.payment_date >= start_date, Payment.payment_date <= end_date))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}
        invoices = {i.invoice_id: i for i in (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"]))).scalars().all()}

        wb = Workbook(); ws = wb.active; ws.title = "Payment Register"
        headers = ["Payment Date", "Invoice No", "Client Name", "GSTIN", "Payment Mode", "Reference No", "Currency", "Currency Amount", "Exch. Rate", "Amount (INR)", "Notes"]
        for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
        for r, p in enumerate(payments, 2):
            c = clients.get(p.client_id, {}); inv = invoices.get(p.invoice_id, {})
            inv_cur = inv.currency if hasattr(inv, 'currency') and inv.currency else 'INR'
            inv_exr = inv.exchange_rate if hasattr(inv, 'exchange_rate') and inv.exchange_rate else 1
            ws.cell(row=r, column=1, value=p.payment_date)
            ws.cell(row=r, column=2, value=inv.invoice_number if hasattr(inv, 'invoice_number') else 'N/A')
            ws.cell(row=r, column=3, value=c.name if hasattr(c, 'name') else 'Unknown')
            ws.cell(row=r, column=4, value=c.gstin if hasattr(c, 'gstin') else '')
            ws.cell(row=r, column=5, value=p.payment_mode); ws.cell(row=r, column=6, value=p.reference_number or '')
            ws.cell(row=r, column=7, value=inv_cur)
            ws.cell(row=r, column=8, value=p.amount)
            ws.cell(row=r, column=9, value=inv_exr)
            ws.cell(row=r, column=10, value=_inr(p.amount, inv_cur, inv_exr))
            ws.cell(row=r, column=11, value=p.notes or '')
        add_totals_row(ws, {8, 10})
        style_header(ws); auto_cols(ws)
        output = BytesIO(); wb.save(output); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition": f"attachment; filename=Payment_Register_{start_date}_{end_date}.xlsx"})
    except Exception as _exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report error: {_exc}")

@router.get("/outstanding-register")
async def get_outstanding_register(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"], Invoice.outstanding > 0))).scalars().all()
    clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}

    wb = Workbook(); ws = wb.active; ws.title = "Outstanding Register"
    headers = ["Invoice No", "Invoice Date", "Due Date", "Client Name", "GSTIN", "Currency", "Invoice Value (Orig.)", "Exch. Rate", "Invoice Value (INR)", "Paid (INR)", "Outstanding (INR)", "Overdue Days", "Status"]
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    today = datetime.now().date()
    for r, inv in enumerate(invoices, 2):
        c = clients.get(inv.client_id)
        try:
            due = datetime.strptime(inv.due_date or '', '%Y-%m-%d').date()
            overdue = (today - due).days if today > due else 0
        except: overdue = 0
        ws.cell(row=r, column=1, value=inv.invoice_number); ws.cell(row=r, column=2, value=inv.invoice_date)
        ws.cell(row=r, column=3, value=inv.due_date)
        ws.cell(row=r, column=4, value=c.name if c else inv.client_name or 'Unknown')
        ws.cell(row=r, column=5, value=c.gstin if c else inv.client_gstin or '')
        cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
        ws.cell(row=r, column=6, value=cur)
        ws.cell(row=r, column=7, value=inv.total or 0)
        ws.cell(row=r, column=8, value=exr)
        ws.cell(row=r, column=9, value=_inr(inv.total or 0, cur, exr))
        ws.cell(row=r, column=10, value=_inr(inv.advance_paid or 0, cur, exr))
        ws.cell(row=r, column=11, value=_inr(inv.outstanding or 0, cur, exr))
        ws.cell(row=r, column=12, value=overdue)
        ws.cell(row=r, column=13, value="Overdue" if overdue > 0 else "Not Due")
    add_totals_row(ws, {7,9,10,11})
    style_header(ws); auto_cols(ws)
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=Outstanding_Register.xlsx"})


@router.get("/sale-register")
async def get_sale_register(start_date: str = Query(...), end_date: str = Query(...),
                              register_type: str = Query("summary"), user: dict = Depends(get_current_user),
                              db: AsyncSession = Depends(get_db)):
    try:
        invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"],
                                                             Invoice.invoice_date >= start_date, Invoice.invoice_date <= end_date))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()}

        wb = Workbook(); ws = wb.active
        if register_type == "summary":
            ws.title = "Sale Register Summary"
            headers = ["Invoice No", "Date", "Client", "GSTIN", "State", "Currency", "Orig. Subtotal", "Exch. Rate", "Subtotal (INR)", "CGST (INR)", "SGST (INR)", "IGST (INR)", "Total (INR)", "Status"]
            for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
            for r, inv in enumerate(invoices, 2):
                c = clients.get(inv.client_id)
                ws.cell(row=r,column=1,value=inv.invoice_number); ws.cell(row=r,column=2,value=inv.invoice_date)
                ws.cell(row=r,column=3,value=c.name if c else inv.client_name or 'Unknown')
                ws.cell(row=r,column=4,value=c.gstin if c else inv.client_gstin or '')
                cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
                ws.cell(row=r,column=5,value=inv.client_state or '')
                ws.cell(row=r,column=6,value=cur); ws.cell(row=r,column=7,value=inv.subtotal or 0)
                ws.cell(row=r,column=8,value=exr)
                ws.cell(row=r,column=9,value=_inr(inv.subtotal or 0, cur, exr))
                ws.cell(row=r,column=10,value=_inr(inv.cgst or 0, cur, exr))
                ws.cell(row=r,column=11,value=_inr(inv.sgst or 0, cur, exr))
                ws.cell(row=r,column=12,value=_inr(inv.igst or 0, cur, exr))
                ws.cell(row=r,column=13,value=_inr(inv.total or 0, cur, exr))
                ws.cell(row=r,column=14,value=inv.status)
        elif register_type in ('itemized', 'items', 'detail', 'item'):
            ws.title = "Sale Register Items"
            headers = ["Invoice No", "Date", "Client", "GSTIN", "State", "Description", "Item Notes", "HSN/SAC", "Qty", "Rate", "GST%", "Amount", "Pure Agent"]
            for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
            r = 2
            for inv in invoices:
                c = clients.get(inv.client_id)
                for item in (inv.items or []):
                    ws.cell(row=r,column=1,value=inv.invoice_number); ws.cell(row=r,column=2,value=inv.invoice_date)
                    ws.cell(row=r,column=3,value=c.name if c else inv.client_name or 'Unknown')
                    ws.cell(row=r,column=4,value=c.gstin if c else inv.client_gstin or '')
                    ws.cell(row=r,column=5,value=inv.client_state or ''); ws.cell(row=r,column=6,value=item.get('description',''))
                    ws.cell(row=r,column=7,value=item.get('item_notes','')); ws.cell(row=r,column=8,value=item.get('hsn_sac_code',''))
                    ws.cell(row=r,column=9,value=item.get('quantity',0)); ws.cell(row=r,column=10,value=item.get('rate',0))
                    ws.cell(row=r,column=11,value=item.get('gst_rate',0)); ws.cell(row=r,column=12,value=item.get('amount',0))
                    ws.cell(row=r,column=13,value="Yes" if item.get('is_pure_agent') else "No"); r += 1

        style_header(ws); auto_cols(ws)

        # Credit Notes sheet
        credit_notes_sr = (await db.execute(select(CreditNote).where(
            CreditNote.tenant_id == user["tenant_id"],
            CreditNote.credit_date >= start_date, CreditNote.credit_date <= end_date
        ))).scalars().all()
        if credit_notes_sr:
            ws_cn = wb.create_sheet("Credit Notes Issued")
            cn_hdrs = ["Credit Note No", "Date", "Client", "GSTIN", "State", "Reason", "Subtotal", "CGST", "SGST", "IGST", "Total"]
            for col, h in enumerate(cn_hdrs, 1): ws_cn.cell(row=1, column=col, value=h)
            for r, cn in enumerate(credit_notes_sr, 2):
                c = clients.get(cn.client_id)
                ws_cn.cell(row=r,column=1,value=cn.credit_note_number); ws_cn.cell(row=r,column=2,value=cn.credit_date)
                ws_cn.cell(row=r,column=3,value=c.name if c else "Unknown"); ws_cn.cell(row=r,column=4,value=c.gstin if c else "")
                ws_cn.cell(row=r,column=5,value=cn.client_state or ""); ws_cn.cell(row=r,column=6,value=cn.reason or "")
                ws_cn.cell(row=r,column=7,value=cn.subtotal or 0); ws_cn.cell(row=r,column=8,value=cn.cgst or 0)
                ws_cn.cell(row=r,column=9,value=cn.sgst or 0); ws_cn.cell(row=r,column=10,value=cn.igst or 0)
                ws_cn.cell(row=r,column=11,value=cn.total or 0)
            add_totals_row(ws_cn, {7,8,9,10,11})
            style_header(ws_cn); auto_cols(ws_cn)

        output = BytesIO(); wb.save(output); output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition": f"attachment; filename=Sale_Register_{register_type}_{start_date}_{end_date}.xlsx"})
    except Exception as _exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report error: {_exc}")

@router.get("/tds-receivable")
async def get_tds_receivable_report(start_date: str = Query(...), end_date: str = Query(...),
                                     user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """
    TDS Receivable report based on actual payment and advance records received in the period.
    Lists each payment/advance with the associated invoice, client, and leaves TDS fields
    blank for manual entry — no automatic 2% calculation.
    """
    try:
        payments = (await db.execute(select(Payment).where(
            Payment.tenant_id == user["tenant_id"],
            Payment.payment_date >= start_date,
            Payment.payment_date <= end_date
        ))).scalars().all()
        advances = (await db.execute(select(Advance).where(
            Advance.tenant_id == user["tenant_id"],
            Advance.payment_date >= start_date,
            Advance.payment_date <= end_date
        ))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(
            select(Client).where(Client.tenant_id == user["tenant_id"])
        )).scalars().all()}
        invoices = {i.invoice_id: i for i in (await db.execute(
            select(Invoice).where(Invoice.tenant_id == user["tenant_id"])
        )).scalars().all()}

        wb = Workbook()
        ws = wb.active
        ws.title = "TDS Receivable"

        headers = [
            "Date", "Type", "Reference / Invoice No", "Client Name", "Client GSTIN",
            "Currency", "Exchange Rate", "Amount (Orig. Currency)", "Amount (INR)", "TDS Rate (%)", "TDS Amount", "Net Received After TDS"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        r = 2
        # Payments
        for p in payments:
            c = clients.get(p.client_id)
            inv = invoices.get(p.invoice_id)
            inv_cur = (inv.currency if inv else None) or 'INR'
            inv_exr = (inv.exchange_rate if inv else None) or 1
            ws.cell(row=r, column=1, value=p.payment_date)
            ws.cell(row=r, column=2, value="Payment")
            ws.cell(row=r, column=3, value=inv.invoice_number if inv else p.reference_number or "")
            ws.cell(row=r, column=4, value=c.name if c else "Unknown")
            ws.cell(row=r, column=5, value=c.gstin if c else "")
            ws.cell(row=r, column=6, value=inv_cur)
            ws.cell(row=r, column=7, value=inv_exr)
            ws.cell(row=r, column=8, value=p.amount or 0)
            ws.cell(row=r, column=9, value=_inr(p.amount or 0, inv_cur, inv_exr))
            ws.cell(row=r, column=10, value="")   # TDS Rate — manual entry
            ws.cell(row=r, column=11, value="")   # TDS Amount — manual entry
            ws.cell(row=r, column=12, value="")   # Net — manual entry
            r += 1

        # Advances
        for adv in advances:
            c = clients.get(adv.client_id)
            adv_cur = adv.currency or 'INR'
            adv_exr = adv.exchange_rate or 1
            ws.cell(row=r, column=1, value=adv.payment_date)
            ws.cell(row=r, column=2, value="Advance")
            ws.cell(row=r, column=3, value=adv.reference_number or "")
            ws.cell(row=r, column=4, value=c.name if c else "Unknown")
            ws.cell(row=r, column=5, value=c.gstin if c else "")
            ws.cell(row=r, column=6, value=adv_cur)
            ws.cell(row=r, column=7, value=adv_exr)
            ws.cell(row=r, column=8, value=adv.amount or 0)
            ws.cell(row=r, column=9, value=_inr(adv.amount or 0, adv_cur, adv_exr))
            ws.cell(row=r, column=10, value="")
            ws.cell(row=r, column=11, value="")
            ws.cell(row=r, column=12, value="")
            r += 1

        # Add note row
        note_row = r + 1
        ws.merge_cells(f"A{note_row}:I{note_row}")
        note = ws.cell(row=note_row, column=1,
                       value="NOTE: Fill in TDS Rate (%) column manually. TDS Amount and Net Received will need to be calculated accordingly.")
        note.font = Font(italic=True, color="DC2626", size=9)

        style_header(ws)
        # Light yellow fill for manual-entry columns
        from openpyxl.styles import PatternFill as PF
        yellow = PF(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")
        for row_idx in range(2, r):
            for col_idx in [10, 11, 12]:
                ws.cell(row=row_idx, column=col_idx).fill = yellow

        auto_cols(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=TDS_Receivable_{start_date}_{end_date}.xlsx"}
        )
    except Exception as exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"TDS report generation failed: {exc}")

@router.get("/complete-report")
async def get_complete_report(start_date: str = Query(...), end_date: str = Query(...),
                               user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint

    business = (await db.execute(select(Business).where(Business.tenant_id == user["tenant_id"]))).scalar_one_or_none()
    company_name = business.business_name if business else "Company"

    invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"],
                                                         Invoice.invoice_date >= start_date, Invoice.invoice_date <= end_date))).scalars().all()
    payments = (await db.execute(select(Payment).where(Payment.tenant_id == user["tenant_id"],
                                                         Payment.payment_date >= start_date, Payment.payment_date <= end_date))).scalars().all()
    credit_notes = (await db.execute(select(CreditNote).where(CreditNote.tenant_id == user["tenant_id"],
                                                               CreditNote.credit_date >= start_date, CreditNote.credit_date <= end_date))).scalars().all()
    estimates = (await db.execute(select(Estimate).where(Estimate.tenant_id == user["tenant_id"],
                                                          Estimate.estimate_date >= start_date, Estimate.estimate_date <= end_date))).scalars().all()
    outstanding_invoices = (await db.execute(select(Invoice).where(Invoice.tenant_id == user["tenant_id"], Invoice.outstanding > 0))).scalars().all()
    advances = (await db.execute(select(Advance).where(Advance.tenant_id == user["tenant_id"],
                                                        Advance.payment_date >= start_date, Advance.payment_date <= end_date))).scalars().all()
    clients_list = (await db.execute(select(Client).where(Client.tenant_id == user["tenant_id"]))).scalars().all()
    clients = {c.client_id: c for c in clients_list}

    today_dt = datetime.now().date()

    # ── Computed summaries — ALL values converted to INR ─────────
    total_revenue    = sum(_inr(i.total or 0,       i.currency or 'INR', i.exchange_rate) for i in invoices if i.status != 'cancelled')
    total_collected  = sum(_inr(p.amount or 0,      (inv_lookup.get(p.invoice_id, {}) or {}).get('currency', 'INR'),
                                (inv_lookup.get(p.invoice_id, {}) or {}).get('exchange_rate', 1)) for p in payments)
    total_outstanding= sum(_inr(i.outstanding or 0, i.currency or 'INR', i.exchange_rate) for i in outstanding_invoices)
    total_cn         = sum(_inr(cn.total or 0,      cn.currency or 'INR', cn.exchange_rate) for cn in credit_notes)
    total_est        = sum(_inr(e.total or 0,       e.currency or 'INR', e.exchange_rate) for e in estimates)
    total_adv        = sum(_inr(a.amount or 0,      a.currency or 'INR', a.exchange_rate) for a in advances)
    paid_inv    = sum(1 for i in invoices if (i.status or '').lower() == 'paid')
    partial_inv = sum(1 for i in invoices if (i.status or '').lower() == 'partial')
    unpaid_inv  = sum(1 for i in invoices if (i.status or '').lower() not in ('paid', 'partial', 'cancelled'))

    # INR invoice lookup for payment conversion
    inv_lookup = {i.invoice_id: {'currency': i.currency or 'INR', 'exchange_rate': i.exchange_rate or 1}
                  for i in invoices}

    # ── Currency-wise INR breakdown for dashboard ─────────────────
    cur_breakdown: dict = {}
    for i in invoices:
        if i.status == 'cancelled': continue
        cur = i.currency or 'INR'
        exr = i.exchange_rate or 1
        if cur not in cur_breakdown:
            cur_breakdown[cur] = {'revenue': 0.0, 'outstanding': 0.0, 'count': 0, 'exr': exr}
        cur_breakdown[cur]['revenue']     += _inr(i.total or 0, cur, exr)
        cur_breakdown[cur]['outstanding'] += _inr(i.outstanding or 0, cur, exr)
        cur_breakdown[cur]['count']       += 1
        cur_breakdown[cur]['exr']          = exr  # last seen rate

    # ── COLOURS ───────────────────────────────────────────────────
    C_DARK   = "0F172A"   # header bg
    C_BLUE   = "1E40AF"   # totals row
    C_TEAL   = "0D9488"   # accent 1
    C_GREEN  = "16A34A"   # accent 2
    C_AMBER  = "D97706"   # accent 3
    C_RED    = "DC2626"   # accent 4
    C_INDIGO = "4338CA"   # accent 5
    C_WHITE  = "FFFFFF"
    C_LGRAY  = "F1F5F9"   # alt row bg
    C_LGRAY2 = "E2E8F0"

    def hdr_style(ws, row=1, color=C_DARK):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(color=C_WHITE, bold=True, size=10)
        bdr  = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[row]:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = bdr
        ws.row_dimensions[row].height = 18

    def alt_row(ws, row, start_col=1):
        fill = PatternFill(start_color=C_LGRAY, end_color=C_LGRAY, fill_type="solid")
        for col in range(start_col, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    def totals(ws, num_cols):
        add_totals_row(ws, num_cols)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — MIS DASHBOARD  (ALL VALUES IN INR)
    # ══════════════════════════════════════════════════════════════
    ws_d = wb.active; ws_d.title = "MIS Dashboard"
    ws_d.sheet_view.showGridLines = False
    for col_letter, w in zip('ABCDEFG', [30, 20, 20, 20, 20, 20, 20]):
        ws_d.column_dimensions[col_letter].width = w

    def _hdr_cell(cell, value, bg=C_DARK, fg=C_WHITE, size=10, bold=True):
        cell.value = value
        cell.font = Font(size=size, bold=bold, color=fg)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    def _val_cell(cell, value, color=C_DARK, align='right', bold=True, size=11, bg=C_LGRAY):
        cell.value = value
        cell.font = Font(size=size, bold=bold, color=color)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # ── Row 1: Title ─────────────────────────────────────────────
    ws_d.merge_cells('A1:G1')
    t = ws_d['A1']
    t.value = f"{company_name}  |  MIS Report (All values in INR ₹)"
    t.font = Font(size=18, bold=True, color=C_WHITE)
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = PatternFill(start_color=C_DARK, end_color=C_DARK, fill_type="solid")
    ws_d.row_dimensions[1].height = 36

    # ── Row 2: Period ─────────────────────────────────────────────
    ws_d.merge_cells('A2:G2')
    t2 = ws_d['A2']
    t2.value = f"Period: {start_date}  to  {end_date}   |   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}   |   Exchange rates as recorded on invoices"
    t2.font = Font(size=9, color="64748B", italic=True)
    t2.alignment = Alignment(horizontal='center')
    t2.fill = PatternFill(start_color=C_LGRAY2, end_color=C_LGRAY2, fill_type="solid")
    ws_d.row_dimensions[2].height = 14

    # ── Row 4: KPI header ─────────────────────────────────────────
    ws_d.row_dimensions[4].height = 18
    kpi_hdrs = ["KPI", "INR Value", "", "KPI", "INR Value"]
    for col, h in enumerate(kpi_hdrs, 1):
        _hdr_cell(ws_d.cell(row=4, column=col), h, bg=C_INDIGO)

    # ── Rows 5-12: KPI tiles (2-column layout) ────────────────────
    kpis_left = [
        ("💰 Total Revenue (INR)",    f"₹{total_revenue:,.2f}",     C_INDIGO),
        ("✅ Total Collected (INR)",   f"₹{total_collected:,.2f}",   C_GREEN),
        ("⏳ Outstanding (INR)",       f"₹{total_outstanding:,.2f}", C_AMBER),
        ("🔴 Credit Notes (INR)",      f"₹{total_cn:,.2f}",          C_RED),
    ]
    kpis_right = [
        ("💳 Advance Received (INR)",  f"₹{total_adv:,.2f}",         C_INDIGO),
        ("📊 Estimates (INR)",         f"₹{total_est:,.2f}",          C_TEAL),
        ("📋 Total Invoices",          f"{len([i for i in invoices if i.status!='cancelled'])}",  C_BLUE),
        ("🏢 Active Clients",          f"{len(clients_list)}",        C_GREEN),
    ]
    for i, ((lbl_l, val_l, clr_l), (lbl_r, val_r, clr_r)) in enumerate(zip(kpis_left, kpis_right)):
        row = 5 + i
        ws_d.row_dimensions[row].height = 22
        _val_cell(ws_d.cell(row=row, column=1), lbl_l, color=clr_l, align='left', size=10)
        _val_cell(ws_d.cell(row=row, column=2), val_l, color=clr_l, size=11)
        ws_d.cell(row=row, column=3).value = ""  # spacer
        _val_cell(ws_d.cell(row=row, column=4), lbl_r, color=clr_r, align='left', size=10)
        _val_cell(ws_d.cell(row=row, column=5), val_r, color=clr_r, size=11)

    # ── Row 14: Currency-wise Breakdown header ────────────────────
    ws_d.row_dimensions[14].height = 18
    cur_cols = ["Currency", "Exch. Rate", "Revenue (INR)", "Collected (INR)", "Outstanding (INR)", "Invoice Count"]
    for col, h in enumerate(cur_cols, 1):
        _hdr_cell(ws_d.cell(row=14, column=col), h, bg=C_TEAL)

    # ── Rows 15+: Currency rows ───────────────────────────────────
    cur_row = 15
    total_rev_inr = total_collected_by_cur = 0.0
    sorted_currencies = sorted(cur_breakdown.keys(), key=lambda c: cur_breakdown[c]['revenue'], reverse=True)
    for cur in sorted_currencies:
        d = cur_breakdown[cur]
        bg = C_LGRAY if cur_row % 2 == 0 else "FFFFFF"
        ws_d.row_dimensions[cur_row].height = 18
        _val_cell(ws_d.cell(row=cur_row, column=1), cur, color=C_BLUE, align='center', bg=bg)
        exr_display = f"1 {cur} = ₹{d['exr']:.4f}" if cur != 'INR' else "INR (base)"
        _val_cell(ws_d.cell(row=cur_row, column=2), exr_display, color="475569", align='center', bold=False, size=9, bg=bg)
        _val_cell(ws_d.cell(row=cur_row, column=3), round(d['revenue'], 2), color=C_INDIGO, bg=bg)
        # Collected for this currency (payments on invoices of this currency)
        collected_cur = sum(_inr(p.amount or 0, cur, d['exr'])
                            for p in payments
                            if inv_lookup.get(p.invoice_id, {}).get('currency', 'INR') == cur)
        _val_cell(ws_d.cell(row=cur_row, column=4), round(collected_cur, 2), color=C_GREEN, bg=bg)
        _val_cell(ws_d.cell(row=cur_row, column=5), round(d['outstanding'], 2), color=C_AMBER, bg=bg)
        _val_cell(ws_d.cell(row=cur_row, column=6), d['count'], color=C_DARK, align='center', bg=bg)
        cur_row += 1

    # Totals row for currency breakdown
    _hdr_cell(ws_d.cell(row=cur_row, column=1), "TOTAL (INR)", bg=C_BLUE)
    _hdr_cell(ws_d.cell(row=cur_row, column=2), "")
    _hdr_cell(ws_d.cell(row=cur_row, column=3), round(total_revenue, 2), bg=C_BLUE)
    _hdr_cell(ws_d.cell(row=cur_row, column=4), round(total_collected, 2), bg=C_BLUE)
    _hdr_cell(ws_d.cell(row=cur_row, column=5), round(total_outstanding, 2), bg=C_BLUE)
    _hdr_cell(ws_d.cell(row=cur_row, column=6), len([i for i in invoices if i.status != 'cancelled']), bg=C_BLUE)
    cur_row += 2  # blank gap

    # ── Chart data tables ─────────────────────────────────────────
    chart_row = cur_row
    # Invoice status table (for pie chart)
    _hdr_cell(ws_d.cell(row=chart_row, column=1), "Invoice Status", bg=C_DARK)
    _hdr_cell(ws_d.cell(row=chart_row, column=2), "Count", bg=C_DARK)
    ws_d.cell(row=chart_row+1, column=1).value = "Paid";    ws_d.cell(row=chart_row+1, column=2).value = paid_inv
    ws_d.cell(row=chart_row+2, column=1).value = "Partial"; ws_d.cell(row=chart_row+2, column=2).value = partial_inv
    ws_d.cell(row=chart_row+3, column=1).value = "Unpaid";  ws_d.cell(row=chart_row+3, column=2).value = unpaid_inv

    # Revenue summary table (for bar chart) — all INR
    _hdr_cell(ws_d.cell(row=chart_row, column=4), "Category", bg=C_DARK)
    _hdr_cell(ws_d.cell(row=chart_row, column=5), "INR Amount", bg=C_DARK)
    ws_d.cell(row=chart_row+1, column=4).value = "Revenue";     ws_d.cell(row=chart_row+1, column=5).value = round(total_revenue, 2)
    ws_d.cell(row=chart_row+2, column=4).value = "Collected";   ws_d.cell(row=chart_row+2, column=5).value = round(total_collected, 2)
    ws_d.cell(row=chart_row+3, column=4).value = "Outstanding"; ws_d.cell(row=chart_row+3, column=5).value = round(total_outstanding, 2)
    ws_d.cell(row=chart_row+4, column=4).value = "Credit Notes";ws_d.cell(row=chart_row+4, column=5).value = round(total_cn, 2)
    ws_d.cell(row=chart_row+5, column=4).value = "Advances";    ws_d.cell(row=chart_row+5, column=5).value = round(total_adv, 2)

    # Pie chart — invoice status
    pie = PieChart(); pie.title = "Invoice Status Breakdown"; pie.style = 10
    pie.width = 12; pie.height = 10
    data_ref = Reference(ws_d, min_col=2, min_row=chart_row, max_row=chart_row+3)
    cats_ref = Reference(ws_d, min_col=1, min_row=chart_row+1, max_row=chart_row+3)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    slices = [DataPoint(idx=0), DataPoint(idx=1), DataPoint(idx=2)]
    slices[0].graphicalProperties.solidFill = C_GREEN
    slices[1].graphicalProperties.solidFill = C_AMBER
    slices[2].graphicalProperties.solidFill = C_RED
    pie.series[0].data_points = slices
    ws_d.add_chart(pie, f"A{chart_row+7}")

    # Bar chart — all INR financials
    bar = BarChart(); bar.type = "col"; bar.title = "Financial Summary (All in INR)"
    bar.style = 10; bar.width = 18; bar.height = 10
    bar.grouping = "clustered"
    bdata = Reference(ws_d, min_col=5, min_row=chart_row, max_row=chart_row+5)
    bcats = Reference(ws_d, min_col=4, min_row=chart_row+1, max_row=chart_row+5)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.series[0].graphicalProperties.solidFill = C_INDIGO
    ws_d.add_chart(bar, f"D{chart_row+7}")

    # ── HELPER: write invoice list sheet ──────────────────────────
    def write_invoice_sheet(ws, inv_list, color=C_DARK):
        hdrs = ["Invoice No","Date","Client","GSTIN","State","Currency","Orig. Amount","Exch. Rate","Subtotal (INR)","CGST (INR)","SGST (INR)","IGST (INR)","Total (INR)","Status"]
        for col, h in enumerate(hdrs, 1): ws.cell(row=1, column=col, value=h)
        for r, inv in enumerate(inv_list, 2):
            c = clients.get(inv.client_id)
            cur = getattr(inv,'currency','INR') or 'INR'
            exr = getattr(inv,'exchange_rate',None) or 1
            ws.cell(row=r,column=1,value=inv.invoice_number); ws.cell(row=r,column=2,value=inv.invoice_date)
            ws.cell(row=r,column=3,value=c.name if c else inv.client_name or 'Unknown')
            ws.cell(row=r,column=4,value=c.gstin if c else getattr(inv,'client_gstin','') or '')
            ws.cell(row=r,column=5,value=getattr(inv,'client_state','') or '')
            ws.cell(row=r,column=6,value=cur)
            ws.cell(row=r,column=7,value=getattr(inv,'total',0) or 0)
            ws.cell(row=r,column=8,value=exr)
            ws.cell(row=r,column=9,value=_inr(getattr(inv,'subtotal',0) or 0, cur, exr))
            ws.cell(row=r,column=10,value=_inr(getattr(inv,'cgst',0) or 0, cur, exr))
            ws.cell(row=r,column=11,value=_inr(getattr(inv,'sgst',0) or 0, cur, exr))
            ws.cell(row=r,column=12,value=_inr(getattr(inv,'igst',0) or 0, cur, exr))
            ws.cell(row=r,column=13,value=_inr(getattr(inv,'total',0) or 0, cur, exr))
            ws.cell(row=r,column=14,value=getattr(inv,'status',''))
            if r % 2 == 0: alt_row(ws, r)
        add_totals_row(ws, {7,9,10,11,12,13})
        hdr_style(ws, color=color); auto_cols(ws)

    def write_items_sheet(ws, doc_list, color=C_DARK):
        hdrs = ["Doc No","Date","Client","GSTIN","Currency","Exch. Rate","Description","Item Notes","HSN/SAC","Qty","Rate","GST%","Orig. Amount","Amount (INR)","Pure Agent"]
        for col, h in enumerate(hdrs, 1): ws.cell(row=1, column=col, value=h)
        r = 2
        for doc in doc_list:
            c = clients.get(doc.client_id)
            cur = getattr(doc,'currency','INR') or 'INR'
            exr = getattr(doc,'exchange_rate',None) or 1
            for item in (doc.items or []):
                ws.cell(row=r,column=1,value=getattr(doc,'invoice_number',None) or getattr(doc,'estimate_number',None) or getattr(doc,'credit_note_number',None))
                ws.cell(row=r,column=2,value=getattr(doc,'invoice_date',None) or getattr(doc,'estimate_date',None) or getattr(doc,'credit_date',None))
                ws.cell(row=r,column=3,value=c.name if c else getattr(doc,'client_name','') or 'Unknown')
                ws.cell(row=r,column=4,value=c.gstin if c else getattr(doc,'client_gstin','') or '')
                ws.cell(row=r,column=5,value=cur)
                ws.cell(row=r,column=6,value=exr)
                ws.cell(row=r,column=7,value=item.get('description',''))
                ws.cell(row=r,column=8,value=item.get('item_notes',''))
                ws.cell(row=r,column=9,value=item.get('hsn_sac_code',''))
                ws.cell(row=r,column=10,value=item.get('quantity',0))
                ws.cell(row=r,column=11,value=item.get('rate',0))
                ws.cell(row=r,column=12,value=item.get('gst_rate',0))
                ws.cell(row=r,column=13,value=item.get('amount',0))
                ws.cell(row=r,column=14,value=_inr(item.get('amount',0), cur, exr))
                ws.cell(row=r,column=15,value="Yes" if item.get('is_pure_agent') else "No")
                if r % 2 == 0: alt_row(ws, r)
                r += 1
        add_totals_row(ws, {10,11,13,14})
        hdr_style(ws, color=color); auto_cols(ws)

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — GSTR-1 B2B
    # ══════════════════════════════════════════════════════════════
    ws_b2b = wb.create_sheet("GSTR-1 B2B")
    b2b_hdrs = ["Invoice No","Invoice Date","Client Name","GSTIN","Place of Supply","Currency","Orig. Amount","Exch. Rate","Taxable (INR)","CGST (INR)","SGST (INR)","IGST (INR)","Total Tax (INR)","Invoice Value (INR)"]
    for col, h in enumerate(b2b_hdrs, 1): ws_b2b.cell(row=1, column=col, value=h)
    r = 2
    for inv in invoices:
        c = clients.get(inv.client_id)
        gstin = c.gstin if c else inv.client_gstin
        if not gstin: continue
        ws_b2b.cell(row=r,column=1,value=inv.invoice_number); ws_b2b.cell(row=r,column=2,value=inv.invoice_date)
        ws_b2b.cell(row=r,column=3,value=c.name if c else inv.client_name or 'Unknown')
        ws_b2b.cell(row=r,column=4,value=gstin); ws_b2b.cell(row=r,column=5,value=inv.client_state or '')
        cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
        ws_b2b.cell(row=r,column=6,value=cur); ws_b2b.cell(row=r,column=7,value=inv.subtotal or 0)
        ws_b2b.cell(row=r,column=8,value=exr)
        ws_b2b.cell(row=r,column=9,value=_inr(inv.subtotal or 0,cur,exr)); ws_b2b.cell(row=r,column=10,value=_inr(inv.cgst or 0,cur,exr))
        ws_b2b.cell(row=r,column=11,value=_inr(inv.sgst or 0,cur,exr)); ws_b2b.cell(row=r,column=12,value=_inr(inv.igst or 0,cur,exr))
        ws_b2b.cell(row=r,column=13,value=_inr((inv.cgst or 0)+(inv.sgst or 0)+(inv.igst or 0),cur,exr))
        ws_b2b.cell(row=r,column=14,value=_inr(inv.total or 0,cur,exr))
        if r % 2 == 0: alt_row(ws_b2b, r)
        r += 1
    add_totals_row(ws_b2b, {7,9,10,11,12,13,14}); hdr_style(ws_b2b, color=C_INDIGO); auto_cols(ws_b2b)

    # ══════════════════════════════════════════════════════════════
    # SHEET 3 — GSTR-1 B2C
    # ══════════════════════════════════════════════════════════════
    ws_b2c = wb.create_sheet("GSTR-1 B2C")
    b2c_hdrs = ["Invoice No","Invoice Date","Client Name","Place of Supply","Currency","Orig. Amount","Exch. Rate","Taxable (INR)","CGST (INR)","SGST (INR)","IGST (INR)","Total Tax (INR)","Invoice Value (INR)"]
    for col, h in enumerate(b2c_hdrs, 1): ws_b2c.cell(row=1, column=col, value=h)
    r = 2
    for inv in invoices:
        c = clients.get(inv.client_id)
        gstin = c.gstin if c else inv.client_gstin
        if gstin: continue
        ws_b2c.cell(row=r,column=1,value=inv.invoice_number); ws_b2c.cell(row=r,column=2,value=inv.invoice_date)
        ws_b2c.cell(row=r,column=3,value=c.name if c else inv.client_name or 'Unknown')
        cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
        ws_b2c.cell(row=r,column=4,value=inv.client_state or '')
        ws_b2c.cell(row=r,column=5,value=cur); ws_b2c.cell(row=r,column=6,value=inv.subtotal or 0)
        ws_b2c.cell(row=r,column=7,value=exr)
        ws_b2c.cell(row=r,column=8,value=_inr(inv.subtotal or 0,cur,exr)); ws_b2c.cell(row=r,column=9,value=_inr(inv.cgst or 0,cur,exr))
        ws_b2c.cell(row=r,column=10,value=_inr(inv.sgst or 0,cur,exr)); ws_b2c.cell(row=r,column=11,value=_inr(inv.igst or 0,cur,exr))
        ws_b2c.cell(row=r,column=12,value=_inr((inv.cgst or 0)+(inv.sgst or 0)+(inv.igst or 0),cur,exr))
        ws_b2c.cell(row=r,column=13,value=_inr(inv.total or 0,cur,exr))
        if r % 2 == 0: alt_row(ws_b2c, r)
        r += 1
    add_totals_row(ws_b2c, {6,8,9,10,11,12,13}); hdr_style(ws_b2c, color=C_TEAL); auto_cols(ws_b2c)

    # ══════════════════════════════════════════════════════════════
    # SHEET 4 — Sale Summary
    # ══════════════════════════════════════════════════════════════
    ws_ss = wb.create_sheet("Sale Summary")
    write_invoice_sheet(ws_ss, invoices, color=C_BLUE)

    # ══════════════════════════════════════════════════════════════
    # SHEET 5 — Sale Item Wise
    # ══════════════════════════════════════════════════════════════
    ws_si = wb.create_sheet("Sale Item Wise")
    write_items_sheet(ws_si, invoices, color=C_DARK)

    # ══════════════════════════════════════════════════════════════
    # SHEET 6 — Payment Register
    # ══════════════════════════════════════════════════════════════
    ws_pr = wb.create_sheet("Payment Register")
    inv_map = {i.invoice_id: i for i in invoices}
    for col, h in enumerate(["Payment Date","Invoice No","Client Name","GSTIN","Mode","Reference","Currency","Orig. Amount","Exch. Rate","Amount (INR)","Notes"], 1):
        ws_pr.cell(row=1, column=col, value=h)
    for r, p in enumerate(payments, 2):
        c = clients.get(p.client_id); inv = inv_map.get(p.invoice_id)
        inv_cur = (inv.currency if inv else None) or 'INR'
        inv_exr = (inv.exchange_rate if inv else None) or 1
        ws_pr.cell(row=r,column=1,value=p.payment_date)
        ws_pr.cell(row=r,column=2,value=inv.invoice_number if inv else 'N/A')
        ws_pr.cell(row=r,column=3,value=c.name if c else 'Unknown')
        ws_pr.cell(row=r,column=4,value=c.gstin if c else '')
        ws_pr.cell(row=r,column=5,value=p.payment_mode); ws_pr.cell(row=r,column=6,value=p.reference_number or '')
        ws_pr.cell(row=r,column=7,value=inv_cur)
        ws_pr.cell(row=r,column=8,value=p.amount)
        ws_pr.cell(row=r,column=9,value=inv_exr)
        ws_pr.cell(row=r,column=10,value=_inr(p.amount, inv_cur, inv_exr))
        ws_pr.cell(row=r,column=11,value=p.notes or '')
        if r % 2 == 0: alt_row(ws_pr, r)
    add_totals_row(ws_pr, {8,10}); hdr_style(ws_pr, color=C_GREEN); auto_cols(ws_pr)

    # ══════════════════════════════════════════════════════════════
    # SHEET 7 — Outstanding Register
    # ══════════════════════════════════════════════════════════════
    ws_or = wb.create_sheet("Outstanding Register")
    for col, h in enumerate(["Invoice No","Invoice Date","Due Date","Client","GSTIN","Currency","Orig. Value","Exch. Rate","Invoice Value (INR)","Paid (INR)","Outstanding (INR)","Overdue Days","Status"], 1):
        ws_or.cell(row=1, column=col, value=h)
    for r, inv in enumerate(outstanding_invoices, 2):
        c = clients.get(inv.client_id)
        try: due = datetime.strptime(inv.due_date or '', '%Y-%m-%d').date(); overdue = max(0,(today_dt-due).days)
        except: overdue = 0
        cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
        ws_or.cell(row=r,column=1,value=inv.invoice_number); ws_or.cell(row=r,column=2,value=inv.invoice_date)
        ws_or.cell(row=r,column=3,value=inv.due_date)
        ws_or.cell(row=r,column=4,value=c.name if c else inv.client_name or 'Unknown')
        ws_or.cell(row=r,column=5,value=c.gstin if c else inv.client_gstin or '')
        ws_or.cell(row=r,column=6,value=cur); ws_or.cell(row=r,column=7,value=inv.total or 0)
        ws_or.cell(row=r,column=8,value=exr)
        ws_or.cell(row=r,column=9,value=_inr(inv.total or 0,cur,exr))
        ws_or.cell(row=r,column=10,value=_inr(inv.advance_paid or 0,cur,exr))
        ws_or.cell(row=r,column=11,value=_inr(inv.outstanding or 0,cur,exr))
        ws_or.cell(row=r,column=12,value=overdue); ws_or.cell(row=r,column=13,value="Overdue" if overdue>0 else "Not Due")
        if r % 2 == 0: alt_row(ws_or, r)
    add_totals_row(ws_or, {7,9,10,11}); hdr_style(ws_or, color=C_AMBER); auto_cols(ws_or)

    # ══════════════════════════════════════════════════════════════
    # SHEET 8 — Client Register
    # ══════════════════════════════════════════════════════════════
    ws_cr = wb.create_sheet("Client Register")
    for col, h in enumerate(["Client Name","GSTIN","PAN","Phone","Email","State","Type","Outstanding"], 1):
        ws_cr.cell(row=1, column=col, value=h)
    for r, cl in enumerate(clients_list, 2):
        cl_outstanding = sum(i.outstanding or 0 for i in outstanding_invoices if i.client_id == cl.client_id)
        ws_cr.cell(row=r,column=1,value=cl.name); ws_cr.cell(row=r,column=2,value=cl.gstin or '')
        ws_cr.cell(row=r,column=3,value=getattr(cl,'pan','') or ''); ws_cr.cell(row=r,column=4,value=cl.phone or '')
        ws_cr.cell(row=r,column=5,value=cl.email or ''); ws_cr.cell(row=r,column=6,value=cl.state or '')
        ws_cr.cell(row=r,column=7,value=cl.type or 'B2B'); ws_cr.cell(row=r,column=8,value=cl_outstanding)
        if r % 2 == 0: alt_row(ws_cr, r)
    add_totals_row(ws_cr, {8}); hdr_style(ws_cr, color=C_TEAL); auto_cols(ws_cr)

    # ══════════════════════════════════════════════════════════════
    # SHEET 9 — Advance Register
    # ══════════════════════════════════════════════════════════════
    ws_ar = wb.create_sheet("Advance Register")
    for col, h in enumerate(["Date","Client","Currency","Orig. Amount","Exch. Rate","Amount (INR)","Mode","Status","Reference"], 1):
        ws_ar.cell(row=1, column=col, value=h)
    for r, adv in enumerate(advances, 2):
        c = clients.get(adv.client_id)
        adv_cur = adv.currency or 'INR'; adv_exr = adv.exchange_rate or 1
        ws_ar.cell(row=r,column=1,value=adv.payment_date)
        ws_ar.cell(row=r,column=2,value=c.name if c else 'Unknown')
        ws_ar.cell(row=r,column=3,value=adv_cur)
        ws_ar.cell(row=r,column=4,value=adv.amount or 0)
        ws_ar.cell(row=r,column=5,value=adv_exr)
        ws_ar.cell(row=r,column=6,value=_inr(adv.amount or 0, adv_cur, adv_exr))
        ws_ar.cell(row=r,column=7,value=adv.payment_mode or '')
        ws_ar.cell(row=r,column=8,value=adv.status or '')
        ws_ar.cell(row=r,column=9,value=adv.reference_number or '')
        if r % 2 == 0: alt_row(ws_ar, r)
    add_totals_row(ws_ar, {4,6}); hdr_style(ws_ar, color=C_INDIGO); auto_cols(ws_ar)

    # ══════════════════════════════════════════════════════════════
    # SHEET 10 — CN Summary
    # ══════════════════════════════════════════════════════════════
    ws_cns = wb.create_sheet("CN Summary")
    for col, h in enumerate(["CN No","Date","Client","GSTIN","State","Currency","Orig. Amount","Exch. Rate","Subtotal (INR)","CGST (INR)","SGST (INR)","IGST (INR)","Total (INR)","Status"], 1):
        ws_cns.cell(row=1, column=col, value=h)
    for r, cn in enumerate(credit_notes, 2):
        c = clients.get(cn.client_id)
        cn_cur = cn.currency or 'INR'; cn_exr = cn.exchange_rate or 1
        ws_cns.cell(row=r,column=1,value=cn.credit_note_number); ws_cns.cell(row=r,column=2,value=cn.credit_date)
        ws_cns.cell(row=r,column=3,value=c.name if c else 'Unknown')
        ws_cns.cell(row=r,column=4,value=c.gstin if c else '')
        ws_cns.cell(row=r,column=5,value=cn.client_state or '')
        ws_cns.cell(row=r,column=6,value=cn_cur); ws_cns.cell(row=r,column=7,value=cn.total or 0)
        ws_cns.cell(row=r,column=8,value=cn_exr)
        ws_cns.cell(row=r,column=9,value=_inr(cn.subtotal or 0,cn_cur,cn_exr))
        ws_cns.cell(row=r,column=10,value=_inr(cn.cgst or 0,cn_cur,cn_exr))
        ws_cns.cell(row=r,column=11,value=_inr(cn.sgst or 0,cn_cur,cn_exr))
        ws_cns.cell(row=r,column=12,value=_inr(cn.igst or 0,cn_cur,cn_exr))
        ws_cns.cell(row=r,column=13,value=_inr(cn.total or 0,cn_cur,cn_exr))
        ws_cns.cell(row=r,column=14,value='issued')
        if r % 2 == 0: alt_row(ws_cns, r)
    add_totals_row(ws_cns, {7,9,10,11,12,13}); hdr_style(ws_cns, color=C_RED); auto_cols(ws_cns)

    # ══════════════════════════════════════════════════════════════
    # SHEET 11 — CN Item Wise
    # ══════════════════════════════════════════════════════════════
    ws_cni = wb.create_sheet("CN Item Wise")
    write_items_sheet(ws_cni, credit_notes, color=C_RED)

    # ══════════════════════════════════════════════════════════════
    # SHEET 12 — Estimate Summary
    # ══════════════════════════════════════════════════════════════
    ws_es = wb.create_sheet("Estimate Summary")
    for col, h in enumerate(["Estimate No","Date","Client","GSTIN","State","Currency","Orig. Amount","Exch. Rate","Subtotal (INR)","CGST (INR)","SGST (INR)","IGST (INR)","Total (INR)","Status"], 1):
        ws_es.cell(row=1, column=col, value=h)
    for r, est in enumerate(estimates, 2):
        c = clients.get(est.client_id)
        est_cur = est.currency or 'INR'; est_exr = est.exchange_rate or 1
        ws_es.cell(row=r,column=1,value=est.estimate_number); ws_es.cell(row=r,column=2,value=est.estimate_date)
        ws_es.cell(row=r,column=3,value=c.name if c else est.client_name or 'Unknown')
        ws_es.cell(row=r,column=4,value=c.gstin if c else est.client_gstin or '')
        ws_es.cell(row=r,column=5,value=est.client_state or '')
        ws_es.cell(row=r,column=6,value=est_cur); ws_es.cell(row=r,column=7,value=est.total or 0)
        ws_es.cell(row=r,column=8,value=est_exr)
        ws_es.cell(row=r,column=9,value=_inr(est.subtotal or 0,est_cur,est_exr))
        ws_es.cell(row=r,column=10,value=_inr(est.cgst or 0,est_cur,est_exr))
        ws_es.cell(row=r,column=11,value=_inr(est.sgst or 0,est_cur,est_exr))
        ws_es.cell(row=r,column=12,value=_inr(est.igst or 0,est_cur,est_exr))
        ws_es.cell(row=r,column=13,value=_inr(est.total or 0,est_cur,est_exr))
        ws_es.cell(row=r,column=14,value=est.status or '')
        if r % 2 == 0: alt_row(ws_es, r)
    add_totals_row(ws_es, {7,9,10,11,12,13}); hdr_style(ws_es, color=C_TEAL); auto_cols(ws_es)

    # ══════════════════════════════════════════════════════════════
    # SHEET 13 — Estimate Item Wise
    # ══════════════════════════════════════════════════════════════
    ws_ei = wb.create_sheet("Estimate Item Wise")
    write_items_sheet(ws_ei, estimates, color=C_TEAL)

    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename=Complete_MIS_Report_{start_date}_{end_date}.xlsx"})


@router.get("/account-register")
async def get_account_register(start_date: str = Query(...), end_date: str = Query(...),
                                user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    try:
        """
        Account Register: one row per invoice, with each item in its own set of columns.
        Columns: Invoice No, Date, Client, GSTIN, State, Diary No, Ref No,
                 [Item N Desc, Item N Notes, Item N Amount, Item N GST%] x N items,
                 CGST, SGST, IGST, Total, Status
        The sheet auto-expands column count based on max items in any invoice.
        """
        invoices = (await db.execute(select(Invoice).where(
            Invoice.tenant_id == user["tenant_id"],
            Invoice.invoice_date >= start_date,
            Invoice.invoice_date <= end_date
        ))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(
            select(Client).where(Client.tenant_id == user["tenant_id"])
        )).scalars().all()}

        # Determine max items across all invoices
        max_items = max((len(inv.items or []) for inv in invoices), default=1)
        max_items = max(max_items, 1)

        wb = Workbook()
        ws = wb.active
        ws.title = "Account Register"

        # Build header row
        base_headers = ["Invoice No", "Date", "Client", "GSTIN", "State", "Currency", "Exch. Rate", "Diary No", "Ref No"]
        item_headers = []
        for n in range(1, max_items + 1):
            item_headers += [
                f"Item {n} Description",
                f"Item {n} Notes",
                f"Item {n} Amount",
                f"Item {n} GST Rate",
            ]
        tail_headers = ["CGST (INR)", "SGST (INR)", "IGST (INR)", "Total (INR)", "Status"]
        headers = base_headers + item_headers + tail_headers

        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, inv in enumerate(invoices, 2):
            c = clients.get(inv.client_id)
            items = inv.items or []

            # Base columns
            cur = inv.currency or 'INR'; exr = inv.exchange_rate or 1
            ws.cell(row=row_idx, column=1, value=inv.invoice_number)
            ws.cell(row=row_idx, column=2, value=inv.invoice_date)
            ws.cell(row=row_idx, column=3, value=c.name if c else inv.client_name or "Unknown")
            ws.cell(row=row_idx, column=4, value=c.gstin if c else inv.client_gstin or "")
            ws.cell(row=row_idx, column=5, value=inv.client_state or "")
            ws.cell(row=row_idx, column=6, value=cur)
            ws.cell(row=row_idx, column=7, value=exr)
            ws.cell(row=row_idx, column=8, value=inv.diary_no or "")
            ws.cell(row=row_idx, column=9, value=inv.ref_no or "")

            # Item columns (pad with blanks if fewer items than max)
            col_offset = len(base_headers) + 1
            for item_n in range(max_items):
                item = items[item_n] if item_n < len(items) else {}
                base_col = col_offset + item_n * 4
                ws.cell(row=row_idx, column=base_col,     value=item.get("description", ""))
                ws.cell(row=row_idx, column=base_col + 1, value=item.get("item_notes", "") or item.get("note", ""))
                ws.cell(row=row_idx, column=base_col + 2, value=item.get("amount", 0) or round((item.get("quantity", 0) or item.get("qty", 0)) * (item.get("rate", 0) or 0), 2))
                ws.cell(row=row_idx, column=base_col + 3, value=item.get("gst_rate", 0))

            # Tail columns (INR converted)
            tail_col = col_offset + max_items * 4
            ws.cell(row=row_idx, column=tail_col,     value=_inr(inv.cgst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 1, value=_inr(inv.sgst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 2, value=_inr(inv.igst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 3, value=_inr(inv.total or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 4, value=inv.status or "")

        style_header(ws)
        auto_cols(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Account_Register_{start_date}_{end_date}.xlsx"}
        )
    except Exception as _exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report error: {_exc}")

@router.get("/credit-note-register")
async def get_credit_note_register(start_date: str = Query(...), end_date: str = Query(...),
                                    user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    try:
        """
        Credit Note Register: one row per credit note, with each item in its own set of columns.
        Same format as Account Register but for credit notes.
        """
        credit_notes = (await db.execute(select(CreditNote).where(
            CreditNote.tenant_id == user["tenant_id"],
            CreditNote.credit_date >= start_date,
            CreditNote.credit_date <= end_date
        ))).scalars().all()
        clients = {c.client_id: c for c in (await db.execute(
            select(Client).where(Client.tenant_id == user["tenant_id"])
        )).scalars().all()}

        max_items = max((len(cn.items or []) for cn in credit_notes), default=1)
        max_items = max(max_items, 1)

        wb = Workbook()
        ws = wb.active
        ws.title = "Credit Note Register"

        base_headers = ["Credit Note No", "Date", "Client", "GSTIN", "State", "Currency", "Exch. Rate", "Diary No", "Ref No", "Reason"]
        item_headers = []
        for n in range(1, max_items + 1):
            item_headers += [f"Item {n} Description", f"Item {n} Notes", f"Item {n} Amount", f"Item {n} GST Rate"]
        tail_headers = ["CGST (INR)", "SGST (INR)", "IGST (INR)", "Total (INR)"]
        headers = base_headers + item_headers + tail_headers

        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        for row_idx, cn in enumerate(credit_notes, 2):
            c = clients.get(cn.client_id)
            items = cn.items or []

            ws.cell(row=row_idx, column=1, value=cn.credit_note_number)
            ws.cell(row=row_idx, column=2, value=cn.credit_date)
            ws.cell(row=row_idx, column=3, value=c.name if c else "Unknown")
            ws.cell(row=row_idx, column=4, value=c.gstin if c else "")
            cur = cn.currency or 'INR'; exr = cn.exchange_rate or 1
            ws.cell(row=row_idx, column=5, value=cn.client_state or "")
            ws.cell(row=row_idx, column=6, value=cur)
            ws.cell(row=row_idx, column=7, value=exr)
            ws.cell(row=row_idx, column=8, value=cn.diary_no or "")
            ws.cell(row=row_idx, column=9, value=cn.ref_no or "")
            ws.cell(row=row_idx, column=10, value=cn.reason or "")

            col_offset = len(base_headers) + 1
            for item_n in range(max_items):
                item = items[item_n] if item_n < len(items) else {}
                base = col_offset + item_n * 4
                ws.cell(row=row_idx, column=base,     value=item.get("description", ""))
                ws.cell(row=row_idx, column=base + 1, value=item.get("item_notes", ""))
                ws.cell(row=row_idx, column=base + 2, value=item.get("amount", 0))
                ws.cell(row=row_idx, column=base + 3, value=item.get("gst_rate", 0))

            tail_col = col_offset + max_items * 4
            ws.cell(row=row_idx, column=tail_col,     value=_inr(cn.cgst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 1, value=_inr(cn.sgst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 2, value=_inr(cn.igst or 0, cur, exr))
            ws.cell(row=row_idx, column=tail_col + 3, value=_inr(cn.total or 0, cur, exr))

            if row_idx % 2 == 0:
                alt_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

        style_header(ws)
        auto_cols(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Credit_Note_Register_{start_date}_{end_date}.xlsx"}
        )
    except Exception as _exc:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Report error: {_exc}")
